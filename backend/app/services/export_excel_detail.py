from __future__ import annotations
"""
Export Excel détaillé des primes NMA.
Feuille "Récap" (tous rôles) + une feuille par rôle.
RCR et SV ont un drill-down : leurs collaborateurs apparaissent sous leur propre ligne.
"""
import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from ..models.bonus import Bonus, BonusPeriod
from ..models.employee import Employee
from ..models.sales import SaleData
from ..models.objective import Objective, Gamme

# ── Couleurs ──────────────────────────────────────────────────────────────────
C_VERT_FONCE  = "1B5E20"
C_VERT_MID    = "2E7D32"
C_VERT_CLAIR  = "E8F5E9"
C_OR          = "F9A825"
C_ORANGE      = "2E7D32"   # vert moyen NMA (quanti)
C_GRIS_MED    = "CFD8DC"
C_BLANC       = "FFFFFF"

MOIS_FR = {
    "01": "Janvier", "02": "Février", "03": "Mars", "04": "Avril",
    "05": "Mai",     "06": "Juin",    "07": "Juillet", "08": "Août",
    "09": "Septembre", "10": "Octobre", "11": "Novembre", "12": "Décembre",
}

CRITERIA_MAP = [
    ("RECOUVREMENT",            "Recouvrement"),
    ("PREVISION",               "Fiabilité prévisions"),
    ("PLANNING_AVANT_01",       "Planning avant 01"),
    ("PORTEFEUILLE_ACHAT",      "Portefeuille achat"),
    ("CLIENTS_CROISSANCE",      "Clients croiss. vs N-1"),
    ("CLIENTS_PATES_CROISSANCE","Clients pâtes croiss."),
    ("CLIENTS_ACTIFS",          "Clients fermes actifs"),
    ("TOP_CLIENTS",             "Top clients croiss."),
    ("CRM_CONFORMITE",          "CRM conformité"),
    ("VISITES_JOURNALIERES",    "Visites journalières"),
    ("VISITES_FERMES",          "Visites fermes"),
    ("KPIS_FERME",              "KPIs ferme"),
    ("ACCOMPAGNEMENT",          "Accompagnement manag."),
    ("RECLAMATIONS",            "Réclamations"),
    ("RAPPORTS_CRM",            "Rapports / CRM"),
]

TOP_ALIASES = {
    "TOP15_CROISSANCE": "TOP_CLIENTS",
    "TOP10_CROISSANCE": "TOP_CLIENTS",
    "TOP10_FERMES":     "TOP_CLIENTS",
    "TOP5_CROISSANCE":  "TOP_CLIENTS",
}

ROLE_LABELS = {
    "COMMERCIAL":   "Commerciaux",
    "RCR":          "RCR",
    "SV":           "Superviseurs",
    "ATC_BV":       "ATC Bétail-Volaille",
    "ATC_FARINE":   "ATC Farine",
    "RCE":          "RCE",
    "RESP_TECH_FP": "Resp. Tech. FP",
    "RESP_TECH_BV": "Resp. Tech. BV",
    "DV":           "Dir. Vente",
    "DCMT":         "DCMT",
}

ROLES_ORDER = ["COMMERCIAL", "RCR", "SV", "ATC_BV", "ATC_FARINE",
               "RCE", "RESP_TECH_FP", "RESP_TECH_BV", "DV", "DCMT"]

ROLE_COLORS = {
    "RCR":          "E8F5E9",
    "SV":           "F3E5F5",
    "COMMERCIAL":   "E3F2FD",
    "ATC_BV":       "FFF8E1",
    "ATC_FARINE":   "FBE9E7",
    "RCE":          "E0F7FA",
    "RESP_TECH_FP": "F9FBE7",
    "RESP_TECH_BV": "FCE4EC",
    "DV":           "EDE7F6",
    "DCMT":         "EFEBE9",
}

# ── Colonnes (module-level pour être partagées) ───────────────────────────────
COL_NOM      = 1
COL_ROLE     = 2
COL_ZONE     = 3
COL_OBJ_FAR  = 4
COL_REA_FAR  = 5
COL_TX_FAR   = 6
COL_OBJ_PAT  = 7
COL_REA_PAT  = 8
COL_TX_PAT   = 9
COL_OBJ_BVF  = 10
COL_REA_BVF  = 11
COL_TX_BVF   = 12
COL_OBJ_TOT  = 13
COL_REA_TOT  = 14
COL_TX_TOT   = 15
COL_CA       = 16
COL_PRIME_FX = 17
COL_COMM_NA  = 18
COL_TOT_QANT = 19
COL_CRIT_START = 20
COL_CRIT_END   = COL_CRIT_START + len(CRITERIA_MAP) - 1
COL_TOT_QUAL   = COL_CRIT_END + 1
COL_TOTAL      = COL_TOT_QUAL + 1
TOTAL_COLS     = COL_TOTAL

COL_WIDTHS = {
    COL_NOM: 22, COL_ROLE: 13, COL_ZONE: 10,
    COL_OBJ_FAR: 9, COL_REA_FAR: 9, COL_TX_FAR: 8,
    COL_OBJ_PAT: 9, COL_REA_PAT: 9, COL_TX_PAT: 8,
    COL_OBJ_BVF: 9, COL_REA_BVF: 9, COL_TX_BVF: 8,
    COL_OBJ_TOT: 9, COL_REA_TOT: 9, COL_TX_TOT: 8,
    COL_CA: 14, COL_PRIME_FX: 11, COL_COMM_NA: 13, COL_TOT_QANT: 13,
    COL_TOT_QUAL: 12, COL_TOTAL: 14,
}


# ── Utilitaires de style ──────────────────────────────────────────────────────
def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)

def _border_thin() -> Border:
    s = Side(style="thin", color="BDBDBD")
    return Border(left=s, right=s, top=s, bottom=s)

def _border_medium() -> Border:
    s = Side(style="medium", color="757575")
    return Border(left=s, right=s, top=s, bottom=s)

def _center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def _fcfa_fmt() -> str:
    return '#,##0" FCFA"'

def _pct_fmt() -> str:
    return '0.0"%"'


# ── En-têtes ──────────────────────────────────────────────────────────────────
def _write_headers(ws, mois_label: str, periode: str):
    def section_header(col_start, col_end, label, color):
        ws.merge_cells(start_row=2, start_column=col_start, end_row=2, end_column=col_end)
        c = ws.cell(2, col_start, label)
        c.font = Font(bold=True, size=10, color=C_BLANC)
        c.fill = _fill(color)
        c.alignment = _center()

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=TOTAL_COLS)
    c = ws.cell(1, 1, f"RAPPORT DÉTAILLÉ DES PRIMES — {mois_label.upper()} {periode[:4]}")
    c.font = Font(bold=True, size=14, color=C_BLANC)
    c.fill = _fill(C_VERT_FONCE)
    c.alignment = _center()
    ws.row_dimensions[1].height = 28

    section_header(COL_NOM, COL_ZONE,      "IDENTIFICATION",                            C_VERT_MID)
    section_header(COL_OBJ_FAR, COL_COMM_NA,"PRIME QUANTITATIVE — DÉTAIL PAR GAMME",   C_ORANGE)
    section_header(COL_TOT_QANT, COL_TOT_QANT, "TOTAL\nQUANTI",                        C_ORANGE)
    section_header(COL_CRIT_START, COL_CRIT_END, "PRIME QUALITATIVE — CRITÈRES DÉTAILLÉS", "1B5E20")
    section_header(COL_TOT_QUAL, COL_TOT_QUAL, "TOTAL\nQUALI",                          "1B5E20")
    section_header(COL_TOTAL, COL_TOTAL,    "TOTAL\nGLOBAL",                            C_VERT_FONCE)
    ws.row_dimensions[2].height = 30

    headers = [
        (COL_NOM,      "Nom"),
        (COL_ROLE,     "Rôle"),
        (COL_ZONE,     "Zone"),
        (COL_OBJ_FAR,  "Obj.\nFarine (T)"),
        (COL_REA_FAR,  "Réal.\nFarine (T)"),
        (COL_TX_FAR,   "Taux\nFarine %"),
        (COL_OBJ_PAT,  "Obj.\nPâtes (T)"),
        (COL_REA_PAT,  "Réal.\nPâtes (T)"),
        (COL_TX_PAT,   "Taux\nPâtes %"),
        (COL_OBJ_BVF,  "Obj.\nBVF (T)"),
        (COL_REA_BVF,  "Réal.\nBVF (T)"),
        (COL_TX_BVF,   "Taux\nBVF %"),
        (COL_OBJ_TOT,  "Obj.\nTotal (T)"),
        (COL_REA_TOT,  "Réal.\nTotal (T)"),
        (COL_TX_TOT,   "Taux\nTotal %"),
        (COL_CA,       "CA Mois M\n(FCFA)"),
        (COL_PRIME_FX, "Prime\nFixe"),
        (COL_COMM_NA,  "Commission\nNA (0.5%)"),
        (COL_TOT_QANT, "TOTAL\nQUANTI"),
    ]
    for code, label in CRITERIA_MAP:
        col = COL_CRIT_START + CRITERIA_MAP.index((code, label))
        headers.append((col, label))
    headers.append((COL_TOT_QUAL, "TOTAL\nQUALI"))
    headers.append((COL_TOTAL,    "TOTAL\nGLOBAL"))

    for col, label in headers:
        c = ws.cell(3, col, label)
        c.font = Font(bold=True, size=9, color="212121")
        c.fill = _fill(C_GRIS_MED)
        c.alignment = _center()
        c.border = _border_thin()
    ws.row_dimensions[3].height = 36

    for col, width in COL_WIDTHS.items():
        ws.column_dimensions[get_column_letter(col)].width = width
    for col in range(COL_CRIT_START, COL_CRIT_END + 1):
        ws.column_dimensions[get_column_letter(col)].width = 11

    ws.freeze_panes = "D4"


# ── Écriture d'une ligne employé ──────────────────────────────────────────────
def _write_emp_row(ws, row: int, bonus: Bonus,
                   vol_by_emp_gamme: dict, ca_by_emp: dict, obj_by_emp_gamme: dict,
                   row_fill: PatternFill | None = None,
                   indent: bool = False):
    emp: Employee = bonus.employee
    role = emp.type_poste.value if emp else ""
    fill = row_fill or _fill(ROLE_COLORS.get(role, C_BLANC))

    detail = bonus.detail_json or {}
    criteria_list = detail.get("criteria", [])
    crit_by_code: dict[str, float] = {}
    for c_item in criteria_list:
        code = c_item.get("code", "")
        mapped = TOP_ALIASES.get(code, code)
        crit_by_code[mapped] = crit_by_code.get(mapped, 0) + float(c_item.get("montant_accorde", 0))

    emp_vols = vol_by_emp_gamme.get(emp.id if emp else -1, {})
    emp_objs = obj_by_emp_gamme.get(emp.id if emp else -1, {})

    def gvol(g): return emp_vols.get(g, 0.0)
    def gobj(g): return emp_objs.get(g, 0.0)
    def gtaux(r, o): return round(r / o * 100, 1) if o > 0 else None

    vol_far = gvol(Gamme.FARINE)
    obj_far = gobj(Gamme.FARINE)
    vol_pat = gvol(Gamme.PATES)
    obj_pat = gobj(Gamme.PATES)
    vol_bvf = gvol(Gamme.BETAIL) + gvol(Gamme.VOLAILLE) + gvol(Gamme.BVF)
    obj_bvf = gobj(Gamme.BETAIL) + gobj(Gamme.VOLAILLE) + gobj(Gamme.BVF)
    vol_tot = float(bonus.volume_realise or 0)
    obj_tot = float(bonus.volume_objectif or 0)

    def write(col, value, fmt=None, bold=False, color=None):
        c = ws.cell(row, col, value)
        c.fill = fill
        c.border = _border_thin()
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.font = Font(bold=bold, size=9, color=color if color else "212121")
        if fmt:
            c.number_format = fmt
        return c

    def write_taux(col, val):
        c = write(col, val, _pct_fmt())
        if val is not None:
            if val >= 115:   c.fill = _fill("A5D6A7")
            elif val >= 100: c.fill = _fill("C8E6C9")
            elif val >= 90:  c.fill = _fill("FFF9C4")
            else:            c.fill = _fill("FFCDD2")
        return c

    # Identification
    nom = f"  ↳ {emp.prenom} {emp.nom}" if indent and emp else (f"{emp.prenom} {emp.nom}" if emp else str(bonus.employee_id))
    nom_cell = write(COL_NOM, nom)
    nom_cell.alignment = Alignment(horizontal="left", vertical="center")
    if indent:
        nom_cell.font = Font(italic=True, size=9, color="37474F")
    write(COL_ROLE, role)
    write(COL_ZONE, emp.region.nom if emp and emp.region else "")

    write(COL_OBJ_FAR, obj_far or None, "0.0")
    write(COL_REA_FAR, vol_far or None, "0.0")
    write_taux(COL_TX_FAR, gtaux(vol_far, obj_far))
    write(COL_OBJ_PAT, obj_pat or None, "0.0")
    write(COL_REA_PAT, vol_pat or None, "0.0")
    write_taux(COL_TX_PAT, gtaux(vol_pat, obj_pat))
    write(COL_OBJ_BVF, obj_bvf or None, "0.0")
    write(COL_REA_BVF, vol_bvf or None, "0.0")
    write_taux(COL_TX_BVF, gtaux(vol_bvf, obj_bvf))
    write(COL_OBJ_TOT, obj_tot or None, "0.0")
    write(COL_REA_TOT, vol_tot or None, "0.0")
    taux_tot = float(bonus.taux_atteinte_global) if bonus.taux_atteinte_global else gtaux(vol_tot, obj_tot)
    write_taux(COL_TX_TOT, taux_tot)

    write(COL_CA, ca_by_emp.get(emp.id if emp else -1, 0) or None, _fcfa_fmt())
    write(COL_PRIME_FX, float(bonus.prime_suivi_fixe) or None, _fcfa_fmt())
    comm = float(bonus.commission_nouvelles_affaires)
    c_comm = write(COL_COMM_NA, comm or None, _fcfa_fmt())
    if comm > 0:
        c_comm.fill = _fill("FFF9C4")

    tot_qant = float(bonus.prime_quantitative) + comm
    c_qant = write(COL_TOT_QANT, tot_qant, _fcfa_fmt(), bold=True)
    c_qant.fill = _fill("FFE0B2")

    for i, (code, _) in enumerate(CRITERIA_MAP):
        col = COL_CRIT_START + i
        val = crit_by_code.get(code)
        c = write(col, val, _fcfa_fmt() if val else None)
        if val and val > 0:
            c.fill = _fill("C8E6C9")
        elif val == 0 and code in {c2 for c2, _ in CRITERIA_MAP if any(
                c2 == item.get("code") or TOP_ALIASES.get(item.get("code")) == c2
                for item in criteria_list)}:
            c.fill = _fill("FFCDD2")

    c_qual = write(COL_TOT_QUAL, float(bonus.prime_qualitative), _fcfa_fmt(), bold=True)
    c_qual.fill = _fill("BBDEFB")

    total_val = float(bonus.total)
    c_tot = write(COL_TOTAL, total_val, _fcfa_fmt(), bold=True)
    c_tot.fill = _fill("A5D6A7") if total_val > 0 else _fill("F5F5F5")
    c_tot.font = Font(bold=True, size=10,
                      color=C_VERT_FONCE if total_val > 0 else "757575")

    ws.row_dimensions[row].height = 18


def _write_drilldown_header(ws, row: int, region_name: str):
    """Ligne d'en-tête avant les collaborateurs d'un superviseur."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=TOTAL_COLS)
    c = ws.cell(row, 1, f"  ↳ Collaborateurs — {region_name}")
    c.font = Font(italic=True, bold=True, size=9, color=C_BLANC)
    c.fill = _fill("43A047")
    c.alignment = Alignment(horizontal="left", vertical="center")
    c.border = _border_thin()
    ws.row_dimensions[row].height = 15


def _write_totals_row(ws, row: int, main_rows: list[int]):
    """Ligne de totaux ne sommant que les lignes 'main_rows' (superviseurs)."""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=COL_COMM_NA)
    c = ws.cell(row, 1, "TOTAUX")
    c.font = Font(bold=True, size=10, color=C_BLANC)
    c.fill = _fill(C_VERT_FONCE)
    c.alignment = _center()

    def sum_formula(col):
        ltr = get_column_letter(col)
        parts = "+".join(f"{ltr}{r}" for r in main_rows)
        return f"={parts}" if parts else "=0"

    for col in [COL_TOT_QANT, COL_TOT_QUAL, COL_TOTAL]:
        c = ws.cell(row, col, sum_formula(col))
        c.font = Font(bold=True, size=10, color=C_BLANC)
        c.fill = _fill(C_VERT_FONCE)
        c.number_format = _fcfa_fmt()
        c.alignment = _center()
        c.border = _border_medium()

    for col in range(COL_CRIT_START, COL_CRIT_END + 1):
        c = ws.cell(row, col, sum_formula(col))
        c.font = Font(bold=True, size=9, color=C_BLANC)
        c.fill = _fill(C_VERT_MID)
        c.number_format = _fcfa_fmt()
        c.alignment = _center()
        c.border = _border_thin()

    ws.row_dimensions[row].height = 22


# ── Construction d'une feuille ────────────────────────────────────────────────
def _fill_sheet(ws, bonuses: list, periode: str, mois_label: str,
                vol_by_emp_gamme: dict, ca_by_emp: dict, obj_by_emp_gamme: dict,
                drilldown_map: dict | None = None,
                add_autofilter: bool = True):
    """
    drilldown_map : {bonus.id: [team Bonus objects]} — si fourni, insère les
    collaborateurs en sous-lignes après chaque superviseur.
    """
    _write_headers(ws, mois_label, periode)
    if add_autofilter:
        ws.auto_filter.ref = f"A3:{get_column_letter(TOTAL_COLS)}3"

    row = 4
    main_rows: list[int] = []   # indices des lignes superviseurs (pour totaux)

    for bonus in bonuses:
        main_rows.append(row)
        _write_emp_row(ws, row, bonus, vol_by_emp_gamme, ca_by_emp, obj_by_emp_gamme)
        row += 1

        if drilldown_map and bonus.id in drilldown_map:
            team = drilldown_map[bonus.id]
            if team:
                emp = bonus.employee
                region_name = emp.region.nom if emp and emp.region else "?"
                _write_drilldown_header(ws, row, region_name)
                row += 1
                for tb in sorted(team, key=lambda b: (b.employee.nom if b.employee else "")):
                    _write_emp_row(ws, row, tb,
                                   vol_by_emp_gamme, ca_by_emp, obj_by_emp_gamme,
                                   row_fill=_fill("EFF8EF"),
                                   indent=True)
                    row += 1
                # Séparateur visuel après le groupe
                ws.row_dimensions[row].height = 6
                row += 1

    _write_totals_row(ws, row, main_rows)


# ── Point d'entrée ────────────────────────────────────────────────────────────
def generate_detail_excel(db: Session, periode: str) -> bytes:
    period = db.query(BonusPeriod).filter(BonusPeriod.periode == periode).first()
    if not period:
        raise ValueError(f"Période {periode} introuvable")

    bonuses = (
        db.query(Bonus)
        .filter(Bonus.period_id == period.id)
        .join(Employee, Bonus.employee_id == Employee.id)
        .order_by(Employee.type_poste, Employee.nom)
        .all()
    )
    emp_ids = [b.employee_id for b in bonuses]

    import sqlalchemy

    # Volumes et CA du mois M
    sales_rows = (
        db.query(SaleData.employee_id, SaleData.gamme,
                 sqlalchemy.func.sum(SaleData.volume).label("vol"),
                 sqlalchemy.func.sum(SaleData.montant_ht).label("ca"))
        .filter(SaleData.periode == periode, SaleData.employee_id.in_(emp_ids))
        .group_by(SaleData.employee_id, SaleData.gamme)
        .all()
    )
    vol_by_emp_gamme: dict = defaultdict(lambda: defaultdict(float))
    ca_by_emp: dict = defaultdict(float)
    for row in sales_rows:
        vol_by_emp_gamme[row.employee_id][row.gamme] = float(row.vol or 0)
        ca_by_emp[row.employee_id] += float(row.ca or 0)

    obj_rows = (
        db.query(Objective.employee_id, Objective.gamme,
                 sqlalchemy.func.sum(Objective.objectif_volume).label("vol"))
        .filter(Objective.periode == periode, Objective.employee_id.in_(emp_ids))
        .group_by(Objective.employee_id, Objective.gamme)
        .all()
    )
    obj_by_emp_gamme: dict = defaultdict(lambda: defaultdict(float))
    for row in obj_rows:
        obj_by_emp_gamme[row.employee_id][row.gamme] = float(row.vol or 0)

    # Regroupement par rôle et par région
    by_role: dict[str, list[Bonus]] = defaultdict(list)
    for b in bonuses:
        role = b.employee.type_poste.value if b.employee else "AUTRE"
        by_role[role].append(b)

    # Commerciaux et ATC regroupés par région (pour drill-down des superviseurs)
    team_roles = {"COMMERCIAL", "ATC_BV", "ATC_FARINE"}
    team_by_region: dict[int, list[Bonus]] = defaultdict(list)
    for b in bonuses:
        emp = b.employee
        if emp and emp.type_poste.value in team_roles and emp.region_id:
            team_by_region[emp.region_id].append(b)

    mois_label = MOIS_FR.get(periode[5:], periode[5:])

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = Workbook()

    # Feuille 1 : Récap (tous rôles)
    ws_recap = wb.active
    ws_recap.title = "Récap"
    _fill_sheet(ws_recap, bonuses, periode, mois_label,
                vol_by_emp_gamme, ca_by_emp, obj_by_emp_gamme,
                add_autofilter=True)

    # Feuilles par rôle
    for role in ROLES_ORDER:
        role_bonuses = by_role.get(role, [])
        if not role_bonuses:
            continue

        sheet_title = ROLE_LABELS.get(role, role)
        ws_role = wb.create_sheet(title=sheet_title)

        drilldown_map = None
        if role in ("RCR", "SV"):
            drilldown_map = {}
            for sup_bonus in role_bonuses:
                emp = sup_bonus.employee
                if emp and emp.region_id:
                    team = [b for b in team_by_region.get(emp.region_id, [])
                            if b.employee_id != emp.id]
                    if team:
                        drilldown_map[sup_bonus.id] = team

        _fill_sheet(ws_role, role_bonuses, periode, mois_label,
                    vol_by_emp_gamme, ca_by_emp, obj_by_emp_gamme,
                    drilldown_map=drilldown_map,
                    add_autofilter=(drilldown_map is None))

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()
