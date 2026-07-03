from __future__ import annotations
import io
import unicodedata
from typing import Optional, List
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy.orm import Session
from ..database import get_db
from ..models.objective import Objective, SalesForecast, Client, ClientPortfolio, Gamme
from ..models.employee import Employee
from ..schemas.objective import (
    ObjectiveCreate, ObjectiveRead, ObjectiveUpdate,
    SalesForecastCreate, SalesForecastRead,
    ClientCreate, ClientRead, PortfolioAssign,
)

router = APIRouter(prefix="/api", tags=["objectives"])

GAMME_IMPORT_COLS = ["FARINE", "BETAIL", "VOLAILLE", "PATES"]


def _norm(s: str) -> str:
    s = unicodedata.normalize("NFD", str(s).upper())
    return "".join(c for c in s if not unicodedata.combining(c)).strip()


@router.get("/objectives", response_model=List[ObjectiveRead])
def list_objectives(
    periode: Optional[str] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(Objective)
    if periode:
        q = q.filter(Objective.periode == periode)
    if employee_id:
        q = q.filter(Objective.employee_id == employee_id)
    return q.all()


@router.post("/objectives", response_model=ObjectiveRead, status_code=201)
def upsert_objective(body: ObjectiveCreate, db: Session = Depends(get_db)):
    existing = (
        db.query(Objective)
        .filter(
            Objective.employee_id == body.employee_id,
            Objective.periode == body.periode,
            Objective.gamme == body.gamme,
        )
        .first()
    )
    if existing:
        existing.objectif_volume = body.objectif_volume
        existing.objectif_ca = body.objectif_ca
        db.commit()
        db.refresh(existing)
        return existing
    obj = Objective(**body.model_dump())
    db.add(obj)
    db.commit()
    db.refresh(obj)
    return obj


@router.delete("/objectives/{objective_id}", status_code=204)
def delete_objective(objective_id: int, db: Session = Depends(get_db)):
    obj = db.query(Objective).filter(Objective.id == objective_id).first()
    if not obj:
        raise HTTPException(status_code=404, detail="Objectif introuvable")
    db.delete(obj)
    db.commit()


@router.get("/forecasts", response_model=List[SalesForecastRead])
def list_forecasts(
    periode: Optional[str] = None,
    employee_id: Optional[int] = None,
    db: Session = Depends(get_db),
):
    q = db.query(SalesForecast)
    if periode:
        q = q.filter(SalesForecast.periode == periode)
    if employee_id:
        q = q.filter(SalesForecast.employee_id == employee_id)
    return q.all()


@router.post("/forecasts", response_model=SalesForecastRead, status_code=201)
def upsert_forecast(body: SalesForecastCreate, db: Session = Depends(get_db)):
    existing = (
        db.query(SalesForecast)
        .filter(
            SalesForecast.employee_id == body.employee_id,
            SalesForecast.periode == body.periode,
            SalesForecast.gamme == body.gamme,
        )
        .first()
    )
    if existing:
        existing.volume_prevu = body.volume_prevu
        existing.ca_prevu = body.ca_prevu
        db.commit()
        db.refresh(existing)
        return existing
    f = SalesForecast(**body.model_dump())
    db.add(f)
    db.commit()
    db.refresh(f)
    return f


@router.get("/clients", response_model=List[ClientRead])
def list_clients(db: Session = Depends(get_db)):
    return db.query(Client).order_by(Client.nom).all()


@router.post("/clients", response_model=ClientRead, status_code=201)
def create_client(body: ClientCreate, db: Session = Depends(get_db)):
    client = Client(**body.model_dump())
    db.add(client)
    db.commit()
    db.refresh(client)
    return client


@router.post("/portfolio", status_code=201)
def assign_portfolio(body: PortfolioAssign, db: Session = Depends(get_db)):
    existing = (
        db.query(ClientPortfolio)
        .filter(
            ClientPortfolio.employee_id == body.employee_id,
            ClientPortfolio.client_id == body.client_id,
            ClientPortfolio.annee == body.annee,
        )
        .first()
    )
    if existing:
        return {"detail": "déjà affecté"}
    p = ClientPortfolio(**body.model_dump())
    db.add(p)
    db.commit()
    return {"detail": "affecté"}


@router.get("/portfolio/{employee_id}/{annee}", response_model=List[ClientRead])
def get_portfolio(employee_id: int, annee: int, db: Session = Depends(get_db)):
    rows = (
        db.query(Client)
        .join(ClientPortfolio)
        .filter(
            ClientPortfolio.employee_id == employee_id,
            ClientPortfolio.annee == annee,
        )
        .all()
    )
    return rows


@router.get("/objectives/template")
def download_objectives_template(db: Session = Depends(get_db)):
    employees = (
        db.query(Employee)
        .filter(Employee.actif == True)
        .order_by(Employee.nom)
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Objectifs"

    hdr_font  = Font(bold=True, color="FFFFFF")
    hdr_fill  = PatternFill("solid", fgColor="1B5E20")
    hdr_align = Alignment(horizontal="center")
    lock_fill = PatternFill("solid", fgColor="F5F5F5")
    lock_font = Font(color="888888")

    headers = ["Nom", "Prénom", "Zone", "Rôle", "FARINE", "BETAIL", "VOLAILLE", "PATES"]
    col_widths = [25, 20, 12, 18, 12, 12, 12, 12]
    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        c = ws.cell(1, col, h)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = hdr_align
        ws.column_dimensions[c.column_letter].width = w

    for row, emp in enumerate(employees, 2):
        region_nom = emp.region.nom if emp.region else ""
        for col, val in enumerate([emp.nom, emp.prenom, region_nom, emp.type_poste], 1):
            c = ws.cell(row, col, val)
            c.fill = lock_fill
            c.font = lock_font
        for col in range(5, 9):
            ws.cell(row, col).number_format = "0.00"

    ws.freeze_panes = "E2"
    ws.auto_filter.ref = f"A1:H{ws.max_row}"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=\"objectifs_template.xlsx\""},
    )


@router.post("/objectives/import")
async def import_objectives(
    file: UploadFile = File(...),
    periode: str = Form(...),
    db: Session = Depends(get_db),
):
    content = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(400, "Fichier Excel invalide")

    ws = wb.active
    raw_headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    headers = [_norm(h) if h else "" for h in raw_headers]

    def _find_col(names: list[str]) -> int | None:
        for name in names:
            try:
                return headers.index(_norm(name)) + 1
            except ValueError:
                pass
        return None

    nom_col    = _find_col(["Nom", "NOM"])
    prenom_col = _find_col(["Prénom", "Prenom", "PRENOM"])
    if not nom_col:
        raise HTTPException(400, "Colonne 'Nom' introuvable dans l'en-tête")

    gamme_cols: dict[str, int] = {}
    for gamme in GAMME_IMPORT_COLS:
        c = _find_col([gamme])
        if c:
            gamme_cols[gamme] = c

    if not gamme_cols:
        raise HTTPException(400, "Aucune colonne gamme trouvée (FARINE, BETAIL, VOLAILLE, PATES)")

    # Build employee lookup: normalized variants → employee
    employees = db.query(Employee).filter(Employee.actif == True).all()
    lookup: dict[str, Employee] = {}
    for emp in employees:
        for key in [
            _norm(f"{emp.nom} {emp.prenom}"),
            _norm(f"{emp.prenom} {emp.nom}"),
            _norm(emp.nom),
        ]:
            lookup[key] = emp

    created = updated = 0
    errors: list[dict] = []

    for row_idx in range(2, ws.max_row + 1):
        nom_val = ws.cell(row_idx, nom_col).value
        if not nom_val:
            continue

        prenom_val = ws.cell(row_idx, prenom_col).value if prenom_col else ""
        nom_n   = _norm(str(nom_val))
        prenom_n = _norm(str(prenom_val)) if prenom_val else ""

        emp = (
            lookup.get(f"{nom_n} {prenom_n}".strip())
            or lookup.get(f"{prenom_n} {nom_n}".strip())
            or lookup.get(nom_n)
        )
        if not emp:
            errors.append({"nom": f"{nom_val} {prenom_val or ''}".strip(), "raison": "Employé introuvable"})
            continue

        for gamme, col in gamme_cols.items():
            raw = ws.cell(row_idx, col).value
            if raw is None or str(raw).strip() == "":
                continue
            try:
                vol = float(str(raw).replace(",", "."))
            except (ValueError, TypeError):
                errors.append({"nom": str(nom_val), "raison": f"Valeur {gamme} invalide : {raw}"})
                continue

            existing = (
                db.query(Objective)
                .filter(
                    Objective.employee_id == emp.id,
                    Objective.periode == periode,
                    Objective.gamme == gamme,
                )
                .first()
            )
            if existing:
                existing.objectif_volume = vol
                updated += 1
            else:
                db.add(Objective(
                    employee_id=emp.id,
                    periode=periode,
                    gamme=gamme,
                    objectif_volume=vol,
                ))
                created += 1

    db.commit()
    return {"created": created, "updated": updated, "errors": errors}
